import openpyxl

def generar_codigo_arm_desde_excel(ruta_archivo):
    wb = openpyxl.load_workbook(ruta_archivo, data_only=True)
    hojas_nivel = [nombre for nombre in wb.sheetnames if nombre.lower().startswith("nivel")]
    hojas_nivel.sort()

    codigo_total = ""

    for hoja in hojas_nivel:
        ws = wb[hoja]
        codigo = []
        codigo.append(f"\n@ ===== {hoja} =====")
        codigo.append("    LDR R0, =0x2000")
        codigo.append("    MOV R3, #4")
        codigo.append("")

        for fila in range(10):
            for col in range(10):
                valor = ws.cell(row=fila+1, column=col+1).value
                if valor == 4:
                    codigo.append(f"    STR R3, [R0, #( {fila}*10 + {col} )*4]")

        codigo_total += "\n".join(codigo) + "\n"

    return codigo_total



ruta = "Niveles.xlsm"
codigo = generar_codigo_arm_desde_excel(ruta)
print(codigo)
